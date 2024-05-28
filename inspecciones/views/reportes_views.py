from django.shortcuts import redirect,render,get_object_or_404
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from ..models import Inspeccion,Componente,Inspector,Vulnerabilidad,Equipo
from .utils import admin_required,render_template
from django.utils import timezone
from django.db.models import Max,Q
from datetime import timedelta
from django.http import JsonResponse,HttpResponse

@login_required
def reporte_probabilidades_semana(request):
    # Obtener el año actual
    años_disponibles = Inspeccion.objects.values_list('fecha__year', flat=True).distinct()
    semanas = range(1, 53)
    año_actual = request.GET.get('año', None)
    año_actual = int(año_actual) if año_actual else None

    # Si no se proporciona un claño, usar el año actual
    if año_actual is None:
        año_actual = timezone.now().year

    # Obtener todos los componentes
    componentes = Componente.objects.all()

    # Crear un diccionario para almacenar los datos del informe
    reporte = {}

    # Obtener el número total de semanas en el año
    num_semanas = 52

    # Iterar sobre cada componente
    for componente in componentes:
        # Crear una entrada en el diccionario de informes para este componente
        reporte[componente] = {}

        # Iterar sobre cada semana del año
        for semana in range(1, num_semanas + 1):
            # Obtener la fecha de inicio y fin de la semana
            fecha_inicio_semana = timezone.make_aware(timezone.datetime(año_actual, 1, 1)) + timedelta(weeks=semana - 1)
            fecha_fin_semana = fecha_inicio_semana + timedelta(days=6)

            # Filtrar inspecciones para este componente y esta semana
            inspecciones = Inspeccion.objects.filter(
                componente=componente,
                fecha__range=(fecha_inicio_semana, fecha_fin_semana)
            ).order_by('-fecha')

            # Obtener la última inspección de esta semana
            ultima_inspeccion = inspecciones.first()

            # Si no hay inspecciones en esta semana, usar la última inspección anterior
            if not ultima_inspeccion:
                inspeccion_anterior = Inspeccion.objects.filter(
                    componente=componente,
                    fecha__lt=fecha_inicio_semana
                ).order_by('-fecha').first()
                if inspeccion_anterior:
                    vulnerabilidad = inspeccion_anterior.vulnerabilidad.valor
                    color = inspeccion_anterior.vulnerabilidad.color
                else:
                    vulnerabilidad = None
                    color = '#000000'
            else:
                vulnerabilidad = ultima_inspeccion.vulnerabilidad.valor
                color = ultima_inspeccion.vulnerabilidad.color


            # Almacenar la vulnerabilidad en el diccionario de informes
            reporte[componente][semana] = {'vulnerabilidad': vulnerabilidad, 'color': color}

    # Renderizar el template con los datos del informe
    return render(request, 'reporte_probabilidades_semanal.html', {'reporte': reporte, 'año_actual': año_actual,'años_disponibles':años_disponibles,'semanas':semanas})




@login_required
def valor_maximo_vulnerabilidad(request):

    return render(request, 'diagramaplanta.html')


def valores_maximos_vulnerabilidad_equipo(request):

    # Obtener la fecha más reciente de cada componente
    fechas_maximas = Inspeccion.objects.values('componente').annotate(max_fecha=Max('fecha'))

    # Filtrar las inspecciones que coincidan con las fechas máximas por componente
    ultimas_inspecciones = Inspeccion.objects.filter(componente__in=fechas_maximas.values('componente'), fecha__in=fechas_maximas.values('max_fecha'))

    # Crear un diccionario para almacenar los valores máximos y colores de vulnerabilidad por equipo
    valores_maximos_vulnerabilidad = {}

    # Iterar sobre las últimas inspecciones de cada componente
    for inspeccion in ultimas_inspecciones:
        equipo_id = inspeccion.componente.equipo.id
        valor_vulnerabilidad = inspeccion.vulnerabilidad.valor
        color_vulnerabilidad = inspeccion.vulnerabilidad.color
        equipo_tag= inspeccion.componente.equipo.tag

        # Actualizar el valor máximo de vulnerabilidad para el equipo
        if equipo_tag not in valores_maximos_vulnerabilidad:
            valores_maximos_vulnerabilidad[equipo_tag] = {'valor': valor_vulnerabilidad, 'color': color_vulnerabilidad}
        else:
            if valor_vulnerabilidad > valores_maximos_vulnerabilidad[equipo_tag]['valor']:
                valores_maximos_vulnerabilidad[equipo_tag] = {'valor': valor_vulnerabilidad, 'color': color_vulnerabilidad}


    return JsonResponse(valores_maximos_vulnerabilidad)


@login_required
def dashboard_view(request):
    inspectores = Inspector.objects.all()
    vulnerabilidades = Vulnerabilidad.objects.all()
    equipos=Equipo.objects.all()
    return render(request, 'reportes/dashboard.html',{
        'inspectores': inspectores,
        'vulnerabilidades': vulnerabilidades,
        'equipos':equipos
    })

from django.utils.dates import MONTHS
from django.db.models import Count
from django.db.models.functions import TruncYear,TruncMonth

@login_required
def get_chart_data(request):
    # Datos de inspecciones por año y mes
    inspecciones_por_mes = (
        Inspeccion.objects
        .annotate(year=TruncYear('fecha'), month=TruncMonth('fecha'))
        .values('year', 'month')
        .annotate(total=Count('id'))
        .order_by('year', 'month')
    )

    # Datos de vulnerabilidades por año y mes
    vulnerabilidades_por_mes = (
        Inspeccion.objects
        .annotate(year=TruncYear('fecha'), month=TruncMonth('fecha'))
        .values('year', 'month', 'vulnerabilidad__nombre')
        .annotate(total=Count('id'))
        .order_by('year', 'month', 'vulnerabilidad__nombre')
    )

    labels = [f"{inspeccion['year'].year}-{inspeccion['month'].month}" for inspeccion in inspecciones_por_mes]
    inspecciones_data = [inspeccion['total'] for inspeccion in inspecciones_por_mes]

    # Crear un diccionario para las vulnerabilidades
    vulnerabilidades_dict = {}
    for v in vulnerabilidades_por_mes:
        key = f"{v['year'].year}-{v['month'].month}"
        if key not in vulnerabilidades_dict:
            vulnerabilidades_dict[key] = {}
        vulnerabilidades_dict[key][v['vulnerabilidad__nombre']] = v['total']

    # Transformar el diccionario en datos para Chart.js
    vulnerabilidades_data = {}
    for key in vulnerabilidades_dict:
        for v in vulnerabilidades_dict[key]:
            if v not in vulnerabilidades_data:
                vulnerabilidades_data[v] = []
            vulnerabilidades_data[v].append({
                "x": key,
                "y": vulnerabilidades_dict[key][v]
            })

    chart_data = {
        "labels": labels,
        "datasets": [
            {
                "label": "Inspecciones",
                "backgroundColor": "rgba(75,192,192,0.4)",
                "borderColor": "rgba(75,192,192,1)",
                "data": inspecciones_data
            }
        ]
    }

    # Definir una lista de colores
    colors = [
        "rgba(255, 99, 132, 0.4)",
        "rgba(54, 162, 235, 0.4)",
        "rgba(255, 206, 86, 0.4)",
        "rgba(75, 192, 192, 0.4)",
        "rgba(153, 102, 255, 0.4)",
        "rgba(255, 159, 64, 0.4)"
    ]
    border_colors = [
        "rgba(255, 99, 132, 1)",
        "rgba(54, 162, 235, 1)",
        "rgba(255, 206, 86, 1)",
        "rgba(75, 192, 192, 1)",
        "rgba(153, 102, 255, 1)",
        "rgba(255, 159, 64, 1)"
    ]

    color_index = 0
    for v in vulnerabilidades_data:
        chart_data['datasets'].append({
            "label": v,
            "data": vulnerabilidades_data[v],
            "backgroundColor": colors[color_index % len(colors)],
            "borderColor": border_colors[color_index % len(border_colors)]
        })
        color_index += 1

    return JsonResponse(chart_data)

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

@login_required
def generate_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    inspector_id = request.GET.get('inspector')
    vulnerabilidad_id = request.GET.get('vulnerabilidad')
    equipo_id = request.GET.get('equipo')

    filters = Q()
    if start_date:
        filters &= Q(fecha__gte=start_date)
    if end_date:
        filters &= Q(fecha__lte=end_date)
    if inspector_id:
        filters &= Q(inspector_id=inspector_id)
    if vulnerabilidad_id:
        filters &= Q(vulnerabilidad_id=vulnerabilidad_id)
    if equipo_id:
        filters &= Q(componente__equipo_id=equipo_id)

    inspecciones = Inspeccion.objects.filter(filters)

    # Crear el libro de Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Reporte de Inspecciones"

    # Encabezados de las columnas
    headers = ['Folio','Fuente de Vulnerabilidad', 'Fecha', 'Ruta Base','Equipo','Componente','Notificacion','Work Order','Probabilidad', 'Condición', 'Recomendación', 'Fecha Planeada', 'Comentario','Realizado']
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar el ancho de las columnas
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 20

    # Agregar los datos de inspecciones
    for row_num, inspeccion in enumerate(inspecciones, 2):
        sheet.cell(row=row_num, column=1, value=inspeccion.id)
        sheet.cell(row=row_num,column=2,value=inspeccion.fuentedevulnerabilidad.nombre)
        sheet.cell(row=row_num, column=3, value=inspeccion.fecha.strftime('%Y-%m-%d %H:%M:%S'))
        sheet.cell(row=row_num, column=4, value=inspeccion.componente.equipo.short_name)
        sheet.cell(row=row_num, column=5, value=inspeccion.componente.equipo.nombre)
        sheet.cell(row=row_num, column=6, value=inspeccion.componente.nombre)
        sheet.cell(row=row_num, column=7, value=inspeccion.notificacion)
        sheet.cell(row=row_num, column=8, value=inspeccion.aviso)
        sheet.cell(row=row_num, column=9, value=inspeccion.vulnerabilidad.nombre)
        sheet.cell(row=row_num,column=10,value=inspeccion.observacion)
        sheet.cell(row=row_num,column=11,value=inspeccion.recomendacion)
        sheet.cell(row=row_num,column=12,value=inspeccion.fechaplaneada.strftime('%Y-%m-%d') if inspeccion.fechaplaneada!=None else '')
        sheet.cell(row=row_num,column=13,value=inspeccion.comentarios)
        sheet.cell(row=row_num,column=14,value=inspeccion.realizado)



    # Guardar el libro de Excel en un buffer
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_inspecciones.xlsx"'
    workbook.save(response)
    
    return response